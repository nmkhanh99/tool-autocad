"""Xuất BOM ra Excel (.xlsx) bằng openpyxl."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .bom import FittingRow, PipeRow
from .layers import LayerRow

_HEAD = Font(bold=True, color="FFFFFF")
_FILL = PatternFill("solid", fgColor="305496")
_MEP_FILL = PatternFill("solid", fgColor="E2EFDA")


def _style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEAD
        cell.fill = _FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 50)


def write_bom(
    path: str | Path,
    fittings: list[FittingRow],
    pipes: list[PipeRow],
) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Phụ kiện"
    ws.append(["Loại phụ kiện", "Đường kính", "Vật liệu", "Số lượng", "Block mẫu"])
    for r in fittings:
        ws.append([r.category, r.dn, r.material, r.count, r.sample_block])
    total = sum(r.count for r in fittings)
    ws.append(["TỔNG", "", "", total, ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    _style_header(ws, 5)
    _autosize(ws)

    ws2 = wb.create_sheet("Ống (theo layer)")
    ws2.append(["Layer", "Hệ MEP?", "Tổng chiều dài (m)"])
    for r in pipes:
        ws2.append([r.layer, "MEP" if r.is_mep else "", r.length_m])
        if r.is_mep:
            for c in range(1, 4):
                ws2.cell(row=ws2.max_row, column=c).fill = _MEP_FILL
    mep_total = round(sum(r.length_m for r in pipes if r.is_mep), 2)
    ws2.append(["TỔNG ống hệ MEP", "", mep_total])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=3).font = Font(bold=True)
    _style_header(ws2, 3)
    _autosize(ws2)

    out = Path(path)
    wb.save(out)
    return out


def write_layer_map(path: str | Path, rows: list[LayerRow]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Layer mapping"
    ws.append(["Layer gốc", "→ Layer chuẩn (đề xuất)", "Nhóm",
               "Số sheet", "Số entity"])
    for r in rows:
        ws.append([r.original, r.target, r.kind, r.sheets, r.entities])
        if r.target == "KHAC":
            for c in range(1, 6):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor="FCE4D6")  # cam nhạt: cần rà tay
    _style_header(ws, 5)
    _autosize(ws)

    out = Path(path)
    wb.save(out)
    return out
